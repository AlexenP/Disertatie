from io import BytesIO

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.models import Property
from app.routers.properties import serialize_property
from app.security import require_roles
from app.services.market_service import get_latest_sector_market_averages

router = APIRouter(prefix="/reports", tags=["reports"])
CAN_EXPORT_REPORTS = ("admin", "agent", "manager", "developer")

MARKET_LABELS = {
    "sub_piata": "Sub piata",
    "la_piata": "La piata",
    "peste_piata": "Peste piata",
}

STATUS_LABELS = {
    "available": "Disponibila",
    "rented": "Inchiriata",
    "occupied": "Ocupata",
    "sold": "Vanduta",
    "inactive": "Inactiva",
}


def to_contains_pattern(value: str) -> str:
    normalized = value.strip().replace("*", "%")
    if not normalized.startswith("%"):
        normalized = f"%{normalized}"
    if not normalized.endswith("%"):
        normalized = f"{normalized}%"
    return normalized


@router.get("/properties/excel")
def export_properties_excel(
    title_query: str | None = Query(default=None),
    address_query: str | None = Query(default=None),
    sector_ids: list[int] | None = Query(default=None),
    property_type_ids: list[int] | None = Query(default=None),
    surface_min: float | None = Query(default=None, ge=0),
    surface_max: float | None = Query(default=None, ge=0),
    price_min: float | None = Query(default=None, ge=0),
    price_max: float | None = Query(default=None, ge=0),
    market_labels: list[str] | None = Query(default=None),
    rent_min: float | None = Query(default=None, ge=0),
    rent_max: float | None = Query(default=None, ge=0),
    statuses: list[str] | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_roles(CAN_EXPORT_REPORTS)),
):
    query = db.query(Property).filter(Property.owner_admin_id == current_user["portfolio_admin_id"])

    if title_query:
        query = query.filter(Property.title.ilike(to_contains_pattern(title_query)))
    if address_query:
        query = query.filter(Property.address.ilike(to_contains_pattern(address_query)))
    if sector_ids:
        query = query.filter(Property.sector_id.in_(sector_ids))
    if property_type_ids:
        query = query.filter(Property.property_type_id.in_(property_type_ids))
    if surface_min is not None:
        query = query.filter(Property.surface_sqm >= surface_min)
    if surface_max is not None:
        query = query.filter(Property.surface_sqm <= surface_max)
    if price_min is not None:
        query = query.filter(Property.price >= price_min)
    if price_max is not None:
        query = query.filter(Property.price <= price_max)
    if rent_min is not None:
        query = query.filter(Property.monthly_rent >= rent_min)
    if rent_max is not None:
        query = query.filter(Property.monthly_rent <= rent_max)
    if statuses:
        query = query.filter(Property.status.in_(statuses))

    properties = query.order_by(Property.id.asc()).all()
    sector_ids = sorted({p.sector_id for p in properties})
    market_averages = get_latest_sector_market_averages(db, sector_ids)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Proprietati"

    headers = [
        "Cod",
        "Titlu",
        "Adresa",
        "Sector",
        "Tip proprietate",
        "Suprafata mp",
        "Pret EUR",
        "Pret mp",
        "Chirie lunara EUR",
        "Status",
        "Pret mediu sector EUR/mp",
        "Diferenta fata de piata %",
        "Clasificare piata",
    ]
    sheet.append(headers)

    for prop in properties:
        item = serialize_property(prop, market_averages.get(prop.sector_id))

        if market_labels and item["market_label"] not in market_labels:
            continue

        sheet.append([
            item["code"],
            item["title"],
            item["address"],
            item["sector_name"] or f"Sector {item['sector_id']}",
            item["property_type_name"] or "",
            item["surface_sqm"],
            item["price"],
            item["price_sqm"],
            item["monthly_rent"],
            STATUS_LABELS.get(item["status"], item["status"]),
            item["market_average_sqm"],
            item["market_difference_percent"],
            MARKET_LABELS.get(item["market_label"], item["market_label"]),
        ])

    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    numeric_formats = {
        "F": "0.00",
        "G": '#,##0.00 "EUR"',
        "H": '#,##0.00 "EUR"',
        "I": '#,##0.00 "EUR"',
        "K": '#,##0.00 "EUR"',
        "L": '0.00"%"',
    }
    for column_letter, number_format in numeric_formats.items():
        for cell in sheet[column_letter][1:]:
            cell.number_format = number_format

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    headers_response = {
        "Content-Disposition": 'attachment; filename="raport_proprietati_geoestate.xlsx"'
    }

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_response,
    )
